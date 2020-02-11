from openpyxl import Workbook
from tqdm import tqdm
import json
import math
import random

wb = Workbook()

# grab the active worksheet
ws = wb.active

'''ws['A1'] = 'RealTeamName'
ws['B1'] = 'KitNumber'
ws['C1'] = 'Name'
ws['D1'] = 'Status'
ws['E1'] = 'Rating'
ws['F1'] = 'PositionName'

with open('PremierLeagueTeams.json', 'r') as f:
    json_data = json.load(f)

teams = []
for i in json_data['api']['teams']:
    # print(i['team_id'], i['name'])
    teams.append(i['name'])

k = 2  # row num
for i in tqdm(teams):
    with open('Athletes by team/'+i+'.json', 'r') as f:
        json_data = json.load(f)

    c = 1
    for j in json_data['api']['players']:
        # k = 0
        if j['season'] == '2018-2019' and j['league'] == 'Premier League' and j['rating'] != None:
            ws['A'+str(k)] = j['team_name']
            ws['B'+str(k)] = c
            ws['C'+str(k)] = j['player_name']
            ws['D'+str(k)] = 1 if random.random() > .1 else 0
            ws['E'+str(k)] = j['rating']
            ws['F'+str(k)] = j['position']

            c += 1
            k += 1'''

'''ws['A1'] = 'GameweekID'
ws['B1'] = 'Date'
ws['C1'] = 'HomeTeam'
ws['D1'] = 'AwayTeam'
ws['E1'] = 'HomeTeamScore'
ws['F1'] = 'AwayTeamScore'

with open('PremierLeagueFixtures.json', 'r') as f:
    json_data = json.load(f)

j = 2
k = 0
for i in tqdm(json_data['api']['fixtures']):
    ws['A'+str(j)] = math.floor(k/10)+1
    ws['B'+str(j)] = i['event_date'][0:10] + ' ' + i['event_date'][11:19]
    ws['C'+str(j)] = i['homeTeam']['team_name']
    ws['D'+str(j)] = i['awayTeam']['team_name']
    ws['E'+str(j)] = 0
    ws['F'+str(j)] = 0
    k += 1
    j += 1


wb.save("sample.xlsx")'''



ws['A1'] = 'Name'

with open('PremierLeagueTeams.json', 'r') as f:
    json_data = json.load(f)

# teams = []
j = 2
for i in json_data['api']['teams']:
    # print(i['team_id'], i['name'])
    # teams.append(i['name'])
    ws['A'+str(j)] = i['name']
    j += 1

wb.save("sample.xlsx")